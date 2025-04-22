import streamlit as st
import pdfplumber
import pandas as pd
import io
import re
import base64
import os
from typing import List, Dict, Any
from openpyxl import load_workbook # .xlsm 読み書きのため

# ----------------------------
# ページ設定（アイコン指定：ブラウザタブ・ブックマーク用）
# ----------------------------
# icon.ico をページアイコンとして使用
st.set_page_config(
    page_title="【数出表】PDF → Excelへの変換",
    layout="centered",
    # ↓↓↓↓↓ この行の "icon.ico" の直後の特殊な空白を削除しました ↓↓↓↓↓
    page_icon="icon.ico" # アイコンファイルのパスを指定 (スクリプトと同じディレクトリにある想定)
    # ↑↑↑↑↑ この行の "icon.ico" の直後の特殊な空白を削除しました ↑↑↑↑↑
)

# ----------------------------
# UIのスタイル設定（洗練されたモダンデザイン - 暖色系背景）
# ----------------------------
st.markdown("""
    <style>
        /* Google FontsのInter, Robotoをインポート */
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600&family=Roboto:wght@300;400;500&display=swap');

        /* アプリ全体の背景とフォント - 薄いオレンジ系の背景 */
        .stApp {
            background: #fff5e6;
            font-family: 'Inter', sans-serif;
        }

        /* タイトル */
        .title {
            font-size: 1.5rem;
            font-weight: 600;
            color: #333;
            margin-bottom: 5px;
        }

        /* サブタイトル */
        .subtitle {
            font-size: 0.9rem;
            color: #666;
            margin-bottom: 25px;
        }

        /* ファイルアップローダーのカスタマイズ */
        [data-testid="stFileUploader"] {
            background: #ffffff;
            border-radius: 10px;
            border: 1px dashed #d0d0d0;
            padding: 30px 20px;
            margin: 20px 0;
        }

        [data-testid="stFileUploader"] label {
            display: none; /* ラベルを非表示に */
        }

        [data-testid="stFileUploader"] section {
            border: none !important;
            background: transparent !important;
        }

        /* ファイル情報カード */
        .file-card {
            background: white;
            border-radius: 8px;
            padding: 12px 16px;
            margin: 15px 0;
            box-shadow: 0 1px 3px rgba(0,0,0,0.08);
            display: flex;
            align-items: center;
            justify-content: space-between;
            border: 1px solid #eaeaea;
        }

        .file-info {
            display: flex;
            align-items: center;
        }

        .file-icon {
            width: 36px;
            height: 36px;
            border-radius: 6px;
            background-color: #f44336; /* PDFアイコン色 */
            display: flex;
            align-items: center;
            justify-content: center;
            margin-right: 12px;
            color: white;
            font-weight: 500;
            font-size: 14px;
        }

        .file-details {
            display: flex;
            flex-direction: column;
        }

        .file-name {
            font-weight: 500;
            color: #333;
            font-size: 0.9rem;
            margin-bottom: 3px;
        }

        .file-meta {
            font-size: 0.75rem;
            color: #888;
        }

        /* ローディングアニメーション */
        .loading-spinner {
            width: 20px;
            height: 20px;
            border: 2px solid rgba(0,0,0,0.1);
            border-radius: 50%;
            border-top-color: #ff9933; /* テーマカラー */
            animation: spin 1s linear infinite;
        }

        /* 完了チェックアイコン */
        .check-icon {
            color: #ff9933; /* テーマカラー */
            font-size: 20px;
        }

        @keyframes spin {
            to { transform: rotate(360deg); }
        }

        /* 進行状況バー（簡易表示） */
        .progress-bar {
            height: 4px;
            background-color: #e0e0e0;
            border-radius: 2px;
            width: 100%;
            margin-top: 10px;
            overflow: hidden; /* はみ出し防止 */
        }

        .progress-value {
            height: 100%;
            background-color: #ff9933; /* テーマカラー */
            border-radius: 2px;
            width: 60%; /* 固定値だが、アニメーションなどで動的に見せても良い */
            transition: width 0.5s ease-in-out;
        }

        /* ダウンロードカード */
        .download-card {
            background: white;
            border-radius: 8px;
            padding: 16px;
            margin: 20px 0;
            box-shadow: 0 2px 5px rgba(0,0,0,0.08);
            display: flex;
            align-items: center;
            justify-content: space-between;
            border: 1px solid #eaeaea;
            transition: all 0.2s ease;
            cursor: pointer;
            text-decoration: none; /* リンクの下線削除 */
            color: inherit; /* 親要素の色を継承 */
        }

        .download-card:hover {
            box-shadow: 0 4px 12px rgba(0,0,0,0.12);
            background-color: #fffaf0; /* ホバー時の背景色 */
            transform: translateY(-2px);
        }

        .download-info {
            display: flex;
            align-items: center;
        }

        .download-icon {
            width: 40px;
            height: 40px;
            border-radius: 8px;
            background-color: #ff9933; /* テーマカラー */
            display: flex;
            align-items: center;
            justify-content: center;
            margin-right: 12px;
            color: white;
            font-weight: 500;
            font-size: 16px;
        }

        .download-details {
            display: flex;
            flex-direction: column;
        }

        .download-name {
            font-weight: 500;
            color: #333;
            font-size: 0.9rem;
            margin-bottom: 3px;
        }

        .download-meta {
            font-size: 0.75rem;
            color: #888;
        }

        /* ダウンロードボタン風の表示 */
        .download-button-imitation { /* ボタンとしての機能は a タグ全体が持つ */
            background-color: #ff9933; /* テーマカラー */
            color: white;
            border: none;
            border-radius: 6px;
            padding: 8px 16px;
            font-size: 0.85rem;
            font-weight: 500;
            transition: background-color 0.2s;
            display: flex;
            align-items: center;
        }
        .download-card:hover .download-button-imitation {
             background-color: #e68a00; /* ホバー時のボタン色 */
        }

        .download-button-icon {
            margin-right: 6px;
        }

        /* Streamlit デフォルトスピナーのテキスト非表示用（必要なら） */
        /* .stSpinner > div > div { visibility: hidden; } */
        /* .stSpinner > div > div::after { content:"処理中..."; visibility: visible; } */

        /* Streamlit コンテナのパディング調整 */
        .css-1544g2n { /* Streamlit version specific */
            padding-top: 2rem;
        }
        .css-18e3th9 { /* Streamlit version specific */
            padding-top: 2rem;
        }

        /* セパレーター線 */
        .separator {
            height: 1px;
            background-color: #ffe0b3; /* テーマカラーに合わせた薄い線 */
            margin: 25px 0;
        }
    </style>
""", unsafe_allow_html=True)

# --- メインコンテナ開始 ---
st.markdown('<div class="main-container">', unsafe_allow_html=True)

# --- タイトルとサブタイトル ---
st.markdown('<div class="title">【数出表】PDF → Excelへの変換</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">PDFの数出表をExcelに変換し、同時に盛り付け札を作成します。</div>', unsafe_allow_html=True)

# ----------------------------
# PDF→Excel変換用の関数群
# ----------------------------
def is_number(text: str) -> bool:
    """文字列が数値（整数）かどうかを判定する"""
    return bool(re.match(r'^\d+$', text.strip()))

def get_line_groups(words: List[Dict[str, Any]], y_tolerance: float = 1.2) -> List[List[Dict[str, Any]]]:
    """y座標に基づいて単語を行ごとにグループ化する"""
    if not words:
        return []
    # y座標（top）でソート
    sorted_words = sorted(words, key=lambda w: w['top'])
    groups = []
    current_group = [sorted_words[0]]
    current_top = sorted_words[0]['top']
    for word in sorted_words[1:]:
        # y座標の差が許容範囲内なら同じ行とみなす
        if abs(word['top'] - current_top) <= y_tolerance:
            current_group.append(word)
        else:
            # 新しい行グループを開始
            groups.append(current_group)
            current_group = [word]
            current_top = word['top']
    groups.append(current_group) # 最後のグループを追加
    return groups

def get_vertical_boundaries(page, tolerance: float = 2) -> List[float]:
    """ページの縦線と単語の左右端から列の境界線を推定する"""
    vertical_lines_x = []
    # 縦線を検出
    for line in page.lines:
        if abs(line['x0'] - line['x1']) < tolerance: # ほぼ垂直な線
            vertical_lines_x.append((line['x0'] + line['x1']) / 2)
    # 重複を除きソート
    vertical_lines_x = sorted(list(set(round(x, 1) for x in vertical_lines_x)))

    words = page.extract_words()
    if not words: # 単語がない場合は線の情報だけ返す
        return vertical_lines_x

    # ページの左右の境界も追加
    left_boundary = min(word['x0'] for word in words)
    right_boundary = max(word['x1'] for word in words)

    # 縦線と左右境界を結合してソート
    boundaries = sorted(list(set([round(left_boundary, 1)] + vertical_lines_x + [round(right_boundary, 1)])))

    # 近すぎる境界線をマージする（任意）
    merged_boundaries = []
    if boundaries:
        merged_boundaries.append(boundaries[0])
        for i in range(1, len(boundaries)):
            if boundaries[i] - merged_boundaries[-1] > tolerance * 2: # ある程度離れていたら追加
                 merged_boundaries.append(boundaries[i])
        # 最後の境界が右端でない場合は追加
        if right_boundary > merged_boundaries[-1] + tolerance * 2 :
             merged_boundaries.append(round(right_boundary, 1))
        boundaries = sorted(list(set(merged_boundaries)))


    return boundaries

def split_line_using_boundaries(sorted_words_in_line: List[Dict[str, Any]], boundaries: List[float]) -> List[str]:
    """境界線に基づいて、一行分の単語をセルテキストに分割する"""
    columns = [""] * (len(boundaries) - 1)
    for word in sorted_words_in_line:
        word_center_x = (word['x0'] + word['x1']) / 2
        for i in range(len(boundaries) - 1):
            left = boundaries[i]
            right = boundaries[i+1]
            # 単語の中心が境界内にあるかチェック
            if left <= word_center_x < right:
                columns[i] += word['text'] + " " # セル内で単語間にスペースを追加
                break # 次の単語へ
    # 各セルの末尾の余分なスペースを削除
    return [col.strip() for col in columns]

def extract_text_with_layout(page) -> List[List[str]]:
    """PDFページからレイアウトを考慮してテキストを行と列に抽出する"""
    # 粗めに単語を抽出（許容誤差を調整）
    words = page.extract_words(x_tolerance=3, y_tolerance=3, keep_blank_chars=False)
    if not words:
        return []

    # 列の境界線を推定
    boundaries = get_vertical_boundaries(page)
    if len(boundaries) < 2: # 境界線が不十分な場合
          # 単純に行ごとにテキストを結合するフォールバック
          lines = page.extract_text(layout=False, x_tolerance=3, y_tolerance=3)
          return [[line] for line in lines.split('\n') if line.strip()]


    # 単語を行ごとにグループ化
    row_groups = get_line_groups(words, y_tolerance=1.5) # 行判定の閾値を調整

    result_rows = []
    for group in row_groups:
        # 行内の単語をx座標でソート
        sorted_group = sorted(group, key=lambda w: w['x0'])
        # 境界線を使ってセルに分割
        columns = split_line_using_boundaries(sorted_group, boundaries)
        # 空行でなければ結果に追加
        if any(cell.strip() for cell in columns):
            result_rows.append(columns)

    return result_rows

def remove_extra_empty_columns(rows: List[List[str]]) -> List[List[str]]:
    """すべての行で完全に空である列を削除する"""
    if not rows:
        return rows
    num_cols = max(len(row) for row in rows) if rows else 0
    if num_cols == 0:
        return rows

    # 各列が空かどうかをチェック
    is_col_empty = [True] * num_cols
    for r, row in enumerate(rows):
        for c in range(len(row)):
            if c < num_cols and row[c].strip():
                is_col_empty[c] = False
        # 行の長さが足りない場合、それ以降の列は空とみなす必要はない

    # 保持する列のインデックス
    keep_indices = [c for c in range(num_cols) if not is_col_empty[c]]

    # 新しい行リストを作成
    new_rows = []
    for row in rows:
        new_row = [row[i] if i < len(row) else "" for i in keep_indices]
        new_rows.append(new_row)

    return new_rows


def format_excel_worksheet(worksheet):
    """xlsxwriter ワークシートの書式設定（列幅・行高さ）"""
    # 注意: この関数は xlsxwriter エンジン使用時にのみ有効
    try:
        worksheet.set_column('A:Z', 15) # 列幅を少し狭く調整
        worksheet.set_default_row(18) # 行高さを少し狭く調整
    except AttributeError:
        # openpyxl の worksheet オブジェクトなど、他のエンジンでは無視
        pass


def post_process_rows(rows: List[List[str]]) -> List[List[str]]:
    """データの後処理: 例として「合計」行の上のセルをクリア"""
    new_rows = [row[:] for row in rows] # リストをコピーして変更
    for i, row in enumerate(new_rows):
        for j, cell in enumerate(row):
            # "合計" という文字が含まれるセルを探す
            if "合計" in str(cell): # 文字列に変換してからチェック
                # そのセルが最初の行でなく、上の行にも同じ列が存在する場合
                if i > 0 and j < len(new_rows[i-1]):
                    # 上の行の同じ列を空白にする
                    new_rows[i-1][j] = ""
    return new_rows

def pdf_to_excel_data(pdf_file) -> pd.DataFrame | None:
    """
    PDFファイルを読み込み、最初のページの表形式データをpandas DataFrameとして返す。
    """
    try:
        with pdfplumber.open(pdf_file) as pdf:
            if not pdf.pages:
                st.warning("PDFにページがありません。")
                return None
            page = pdf.pages[0] # 最初のページのみ対象

            # レイアウトを考慮してテキスト抽出
            rows = extract_text_with_layout(page)

            # 空行を除去
            rows = [row for row in rows if any(cell.strip() for cell in row)]
            if not rows:
                st.warning("PDFの最初のページからテキストデータを抽出できませんでした。")
                return None

            # データ後処理（例：「合計」行の上のセルをクリア）
            rows = post_process_rows(rows)

            # 完全に空の列を削除
            rows = remove_extra_empty_columns(rows)
            if not rows or not rows[0]: # 空の列削除後に行や列がなくなった場合
                 st.warning("空の列を削除した結果、データがなくなりました。")
                 return None

            # 最大列数を取得
            max_cols = max(len(row) for row in rows) if rows else 0

            # すべての行が同じ列数になるように空白で埋める
            normalized_rows = [row + [''] * (max_cols - len(row)) for row in rows]

            # DataFrameに変換（ヘッダーなし）
            df = pd.DataFrame(normalized_rows)
            return df

    except Exception as e:
        st.error(f"PDF処理中にエラーが発生しました: {e}")
        return None


# ----------------------------
# テンプレートExcelファイルのパス設定と存在確認
# ----------------------------
template_path = "template.xlsm" # ここで .xlsm を指定
template_wb = None # Workbookオブジェクトを格納する変数

if not os.path.exists(template_path):
    st.error(f"テンプレートファイル '{template_path}' が見つかりません。スクリプトと同じ場所に配置してください。")
    st.stop() # テンプレートがない場合は処理を停止
else:
    # テンプレートファイルの読み込みを試みる
    try:
        # マクロを保持して読み込む
        template_wb = load_workbook(template_path, keep_vba=True)
        # --- st.success(...) の行を削除 ---
    except Exception as e:
        st.error(f"テンプレートファイル '{template_path}' の読み込み中にエラーが発生しました: {e}")
        template_wb = None # エラー時は None に設定
        st.stop() # 読み込めない場合は停止

# ----------------------------
# UI：PDFファイルアップロード
# ----------------------------
uploaded_pdf = st.file_uploader("処理するPDFファイルをアップロードしてください", type="pdf",
                                help="ここにPDFファイルをドラッグ＆ドロップするか、クリックして選択してください。")

# --- ファイル処理とダウンロード表示用のコンテナ ---
file_container = st.container()
download_container = st.container()

# --- PDFがアップロードされたら処理を実行 ---
if uploaded_pdf is not None and template_wb is not None:
    # --- 処理中の表示 ---
    with file_container:
        file_ext = uploaded_pdf.name.split('.')[-1].lower()
        file_icon = "PDF" # PDF固定
        file_size = len(uploaded_pdf.getvalue()) / 1024

        progress_placeholder = st.empty()
        progress_placeholder.markdown(f"""
        <div class="file-card">
            <div class="file-info">
                <div class="file-icon">{file_icon}</div>
                <div class="file-details">
                    <div class="file-name">{uploaded_pdf.name}</div>
                    <div class="file-meta">{file_size:.0f} KB</div>
                </div>
            </div>
            <div class="loading-spinner"></div>
        </div>
        <div class="progress-bar"><div class="progress-value"></div></div>
        """, unsafe_allow_html=True)

    # --- PDFからDataFrameへの変換 ---
    with st.spinner("PDFデータを抽出中..."):
        df_pdf = pdf_to_excel_data(uploaded_pdf)

    if df_pdf is not None and not df_pdf.empty:
        try:
            # --- テンプレートにデータを書き込み ---
            with st.spinner("テンプレートにデータを書き込み中..."):
                # テンプレートの最初のワークシートを取得
                template_ws = template_wb.worksheets[0]

                # 注意: 既存のデータをクリアする場合はここで行う
                # 例: template_ws.delete_rows(1, template_ws.max_row) # 全行削除

                # DataFrameのデータをシートに書き込む (1行目、1列目から)
                for r_idx, row in df_pdf.iterrows():
                    for c_idx, value in enumerate(row):
                        # openpyxlは1始まりのインデックス
                        template_ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)

            # --- メモリ上でExcelファイルを生成 ---
            with st.spinner("Excelファイルを生成中..."):
                output = io.BytesIO()
                # keep_vba=True でロードしたので、そのまま保存すればマクロは保持される
                template_wb.save(output)
                output.seek(0)
                final_excel_bytes = output.read()

            # --- 処理完了表示 ---
            with file_container:
                 progress_placeholder.markdown(f"""
                 <div class="file-card">
                     <div class="file-info">
                         <div class="file-icon">{file_icon}</div>
                         <div class="file-details">
                             <div class="file-name">{uploaded_pdf.name}</div>
                             <div class="file-meta">{file_size:.0f} KB - 処理完了</div>
                         </div>
                     </div>
                     <div class="check-icon">✓</div>
                 </div>
                 """, unsafe_allow_html=True)

            # --- ダウンロードリンクの生成 ---
            with download_container:
                st.markdown('<div class="separator"></div>', unsafe_allow_html=True) # 区切り線

                original_pdf_name = os.path.splitext(uploaded_pdf.name)[0]
                # 出力ファイル名を .xlsm に
                output_filename = f"{original_pdf_name}_Merged.xlsm"
                excel_size = len(final_excel_bytes) / 1024
                b64 = base64.b64encode(final_excel_bytes).decode('utf-8')

                # MIMEタイプを .xlsm 用に設定
                mime_type = "application/vnd.ms-excel.sheet.macroEnabled.12"

                # ダウンロードリンク (HTMLコメント削除済み)
                href = f"""
                <a href="data:{mime_type};base64,{b64}" download="{output_filename}" class="download-card">
                    <div class="download-info">
                        <div class="download-icon">XLSM</div>
                        <div class="download-details">
                            <div class="download-name">{output_filename}</div>
                            <div class="download-meta">Excel (マクロ有効)・{excel_size:.0f} KB</div>
                        </div>
                    </div>
                    <div class="download-button-imitation">
                        <span class="download-button-icon">↓</span>
                        Download
                    </div>
                </a>
                """
                st.markdown(href, unsafe_allow_html=True)

        except Exception as e:
            st.error(f"Excelファイルへの書き込みまたは生成中にエラーが発生しました: {e}")
            # エラー発生時は完了表示を元に戻すか、エラー表示を維持
            with file_container:
                 progress_placeholder.markdown(f"""
                 <div class="file-card" style="border-color: red;">
                     <div class="file-info">
                         <div class="file-icon" style="background-color: red;">!</div>
                         <div class="file-details">
                             <div class="file-name">{uploaded_pdf.name}</div>
                             <div class="file-meta" style="color: red;">処理中にエラーが発生しました</div>
                         </div>
                     </div>
                 </div>
                 """, unsafe_allow_html=True)

    elif df_pdf is None:
        # pdf_to_excel_data 関数内でエラーまたは警告が出力されているはず
        with file_container:
             progress_placeholder.markdown(f"""
             <div class="file-card" style="border-color: orange;">
                 <div class="file-info">
                     <div class="file-icon" style="background-color: orange;">!</div>
                     <div class="file-details">
                         <div class="file-name">{uploaded_pdf.name}</div>
                         <div class="file-meta" style="color: orange;">PDFからデータを抽出できませんでした</div>
                     </div>
                 </div>
             </div>
             """, unsafe_allow_html=True)

# --- テンプレートファイルが見つからないか読み込めなかった場合 ---
elif uploaded_pdf is not None and template_wb is None:
    st.warning("テンプレートファイルが正しく読み込めていないため、処理を開始できません。")


# --- メインコンテナ終了 ---
st.markdown('</div>', unsafe_allow_html=True)
