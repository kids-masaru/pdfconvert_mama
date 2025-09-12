# streamlit_app.py (最終版)

import streamlit as st
import pandas as pd
import io
import os
import re
from openpyxl import load_workbook
import glob

# 正常な `pdf_utils` から `match_bento_data` をインポート
from pdf_utils import (
    safe_write_df, pdf_to_excel_data_for_paste_sheet, extract_table_from_pdf_for_bento,
    find_correct_anchor_for_bento, extract_bento_range_for_bento, match_bento_data, 
    extract_detailed_client_info_from_pdf, export_detailed_client_data_to_dataframe
)

st.set_page_config(
    page_title="PDF変換ツール",
    page_icon="./static/icons/android-chrome-192.png",
    layout="centered",
)

def load_master_data(file_prefix, default_columns):
    """
    CSVファイルを読み込む。
    【重要】1行目をヘッダーとして正しく読み込む。
    """
    list_of_files = glob.glob(os.path.join('.', f'{file_prefix}*.csv'))
    if not list_of_files:
        return pd.DataFrame(columns=default_columns)
    latest_file = max(list_of_files, key=os.path.getmtime)
    encodings = ['utf-8-sig', 'utf-8', 'cp932', 'shift_jis']
    for encoding in encodings:
        try:
            # --- ▼修正点：header=Noneを削除し、正しくヘッダーを読み込む ---
            df = pd.read_csv(latest_file, encoding=encoding, dtype=str).fillna('')
            # --- ▲修正点▲ ---
            if not df.empty: return df
        except Exception:
            continue
    return pd.DataFrame(columns=default_columns)

if 'master_df' not in st.session_state:
    st.session_state.master_df = load_master_data("商品マスタ一覧", ['商品予定名', 'パン箱入数', '商品名', 'クラス分け名称4', 'クラス分け名称5'])
if 'customer_master_df' not in st.session_state:
    st.session_state.customer_master_df = load_master_data("得意先マスタ一覧", ['得意先ＣＤ', '得意先名'])

st.markdown("""
    <style>
        [data-testid="stSidebarNav"] ul { display: none; }
        .custom-title {
            font-size: 2.1rem; font-weight: 600; color: #3A322E;
            padding-bottom: 10px; border-bottom: 3px solid #FF9933; margin-bottom: 25px;
        }
        .stApp { background: #fff5e6; }
    </style>
""", unsafe_allow_html=True)

st.sidebar.title("メニュー")
st.sidebar.page_link("streamlit_app.py", label="PDF Excel 変換", icon="📄")
st.sidebar.page_link("pages/マスタ設定.py", label="マスタ設定", icon="⚙️")
st.markdown('<p class="custom-title">数出表 PDF変換ツール</p>', unsafe_allow_html=True)
show_debug = st.sidebar.checkbox("デバッグ情報を表示", value=False)
uploaded_pdf = st.file_uploader("処理するPDFファイルをアップロードしてください", type="pdf", label_visibility="collapsed")

if uploaded_pdf is not None:
    template_path = "template.xlsm"
    nouhinsyo_path = "nouhinsyo.xlsx"
    if not os.path.exists(template_path) or not os.path.exists(nouhinsyo_path):
        st.error(f"必要なテンプレートファイルが見つかりません：'{template_path}' または '{nouhinsyo_path}'")
        st.stop()
    
    template_wb = load_workbook(template_path, keep_vba=True)
    nouhinsyo_wb = load_workbook(nouhinsyo_path)
    pdf_bytes_io = io.BytesIO(uploaded_pdf.getvalue())
    
    df_paste_sheet, df_bento_sheet, df_client_sheet = None, None, None
    with st.spinner("PDFからデータを抽出中..."):
        try:
            df_paste_sheet = pdf_to_excel_data_for_paste_sheet(io.BytesIO(pdf_bytes_io.getvalue()))
        except Exception as e:
            df_paste_sheet = None
            st.error(f"PDFからの貼り付け用データ抽出中にエラーが発生しました: {str(e)}")

        if df_paste_sheet is not None:
            try:
                tables = extract_table_from_pdf_for_bento(io.BytesIO(pdf_bytes_io.getvalue()))
                if tables:
                    main_table = max(tables, key=len)
                    anchor_col = find_correct_anchor_for_bento(main_table)
                    if anchor_col != -1:
                        bento_list = extract_bento_range_for_bento(main_table, anchor_col)
                        if bento_list:
                            matched_data = match_bento_data(bento_list, st.session_state.master_df)
                            df_bento_sheet = pd.DataFrame(matched_data, columns=['商品予定名', 'パン箱入数', 'クラス分け名称4', 'クラス分け名称5'])
                            
                            if show_debug:
                                st.write("--- 抽出・マッチング後の最終データ ---")
                                st.dataframe(df_bento_sheet)

            except Exception as e:
                st.error(f"注文弁当データ処理中にエラーが発生しました: {str(e)}")
                if show_debug: st.exception(e)

            try:
                client_data = extract_detailed_client_info_from_pdf(io.BytesIO(pdf_bytes_io.getvalue()))
                if client_data:
                    df_client_sheet = export_detailed_client_data_to_dataframe(client_data)
            except Exception as e:
                st.error(f"クライアント情報抽出中にエラーが発生しました: {str(e)}")
    
    if df_paste_sheet is not None:
        try:
            with st.spinner("Excelファイルを作成中..."):
                ws_paste = template_wb["貼り付け用"]
                for r_idx, row in df_paste_sheet.iterrows():
                    for c_idx, value in enumerate(row):
                        ws_paste.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                if df_bento_sheet is not None:
                    safe_write_df(template_wb["注文弁当の抽出"], df_bento_sheet, start_row=1)
                if df_client_sheet is not None:
                    safe_write_df(template_wb["クライアント抽出"], df_client_sheet, start_row=1)
                
                output_macro = io.BytesIO()
                template_wb.save(output_macro)
                macro_excel_bytes = output_macro.getvalue()

                df_bento_for_nouhin = None
                if df_bento_sheet is not None:
                    master_df = st.session_state.master_df.copy()
                    master_df.columns = master_df.columns.str.strip()
                    if not master_df.empty and '商品名' in master_df.columns:
                        master_map = master_df.drop_duplicates(subset=['商品予定名']).set_index('商品予定名')['商品名'].to_dict()
                        df_bento_for_nouhin = df_bento_sheet.copy()
                        df_bento_for_nouhin['商品名'] = df_bento_for_nouhin['商品予定名'].map(master_map)
                        df_bento_for_nouhin = df_bento_for_nouhin[['商品予定名', 'パン箱入数', '商品名']]
                
                ws_paste_n = nouhinsyo_wb["貼り付け用"]
                for r_idx, row in df_paste_sheet.iterrows():
                    for c_idx, value in enumerate(row):
                        ws_paste_n.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                if df_bento_for_nouhin is not None:
                    safe_write_df(nouhinsyo_wb["注文弁当の抽出"], df_bento_for_nouhin, start_row=1)
                if df_client_sheet is not None:
                    safe_write_df(nouhinsyo_wb["クライアント抽出"], df_client_sheet, start_row=1)
                if not st.session_state.customer_master_df.empty:
                    safe_write_df(nouhinsyo_wb["得意先マスタ"], st.session_state.customer_master_df, start_row=1)
                
                output_data_only = io.BytesIO()
                nouhinsyo_wb.save(output_data_only)
                data_only_excel_bytes = output_data_only.getvalue()

            st.success("✅ ファイルの準備が完了しました！")
            original_pdf_name = os.path.splitext(uploaded_pdf.name)[0]
            
            col1, col2 = st.columns(2)
            with col1:
                st.download_button(
                    label="▼　数出表ダウンロード", data=macro_excel_bytes,
                    file_name=f"{original_pdf_name}_数出表.xlsm",
                    mime="application/vnd.ms-excel.sheet.macroEnabled.12"
                )
            with col2:
                st.download_button(
                    label="▼　納品書ダウンロード", data=data_only_excel_bytes,
                    file_name=f"{original_pdf_name}_納品書.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"Excelファイル生成中にエラーが発生しました: {str(e)}")
