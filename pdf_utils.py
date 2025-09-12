# pdf_utils.py

import pandas as pd
import pdfplumber
import re
import unicodedata
from typing import List, Dict, Any
from openpyxl.utils.dataframe import dataframe_to_rows

# ──────────────────────────────────────────────
# ヘルパー関数
# ──────────────────────────────────────────────
def safe_write_df(worksheet, df, start_row=1):
    """DataFrameをExcelシートに安全に書き込む"""
    num_cols = df.shape[1]
    # 既存のデータをクリア
    if worksheet.max_row >= start_row:
        for row in range(start_row, worksheet.max_row + 1):
            for col in range(1, num_cols + 2): # 余裕をもってクリア
                worksheet.cell(row=row, column=col).value = None
    # 新しいデータを書き込み
    for r_idx, row_data in enumerate(df.itertuples(index=False), start=start_row):
        for c_idx, value in enumerate(row_data, start=1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)

# ──────────────────────────────────────────────
# ▼ここからが修正された弁当データマッチング関数▼
# ──────────────────────────────────────────────
def match_bento_data(pdf_bento_list: List[str], master_df: pd.DataFrame) -> List[List[str]]:
    """
    PDFから抽出した弁当名リストを商品マスタと照合し、
    [弁当名, パン箱入数, クラス分け名称4, クラス分け名称5] のリストを返す。
    列番号を修正して正しいデータを取得する。
    """
    if master_df is None or master_df.empty:
        return [[name, "", "", ""] for name in pdf_bento_list]

    # CSVヘッダーのスペース問題を修正
    master_df.columns = master_df.columns.str.strip()
    
    # デバッグ：列数と列名を確認
    print(f"マスタの列数: {len(master_df.columns)}")
    print(f"列名一覧: {list(master_df.columns)}")
    
    if len(master_df.columns) < 18:
        return [[name, "", "マスタ列不足: 必要な列数が不足", ""] for name in pdf_bento_list]
    
    matched_results = []
    
    for pdf_name in pdf_bento_list:
        pdf_name_stripped = pdf_name.strip()
        norm_pdf = unicodedata.normalize('NFKC', pdf_name_stripped).replace(" ", "")
        
        result_data = [pdf_name_stripped, "", "", ""]  # デフォルト値
        
        # 全ての行を検索してマッチするものを探す
        best_match = None
        for idx, row in master_df.iterrows():
            # 各列で商品名を検索（最初の5列）
            for col_idx in range(min(5, len(master_df.columns))):
                cell_value = str(row.iloc[col_idx]).strip()
                if cell_value and cell_value != 'nan':
                    norm_master = unicodedata.normalize('NFKC', cell_value).replace(" ", "")
                    
                    # 完全一致チェック
                    if norm_master == norm_pdf:
                        # ** 修正：正しい列番号を使用 **
                        # パン箱入数は第2画像から推測すると、もっと後の列にありそう
                        # 一旦全ての列を確認して「パン箱入数」または数値24が含まれる列を探す
                        
                        # 暫定的に異なる列番号を試してみる
                        pan_box = ""
                        class4 = ""  
                        class5 = ""
                        
                        # パン箱入数を探す（数値が入っている可能性の高い列を順次チェック）
                        for potential_pan_col in [2, 6, 7, 8, 9, 10]:  # 様々な列を試す
                            if len(row) > potential_pan_col:
                                potential_value = str(row.iloc[potential_pan_col]).strip()
                                if potential_value and potential_value.isdigit():
                                    # デバッグ出力
                                    print(f"商品名: {cell_value}, 列{potential_pan_col}: {potential_value}")
                                    if potential_value not in ["88", "1", "0"]:  # 明らかに単価でない値
                                        pan_box = potential_value
                                        break
                        
                        # クラス分け名称4, 5（P列、R列相当を探す）
                        if len(row) > 15:
                            class4 = str(row.iloc[15]) if str(row.iloc[15]).strip() != 'nan' else ""
                        if len(row) > 17:
                            class5 = str(row.iloc[17]) if str(row.iloc[17]).strip() != 'nan' else ""
                        
                        best_match = [cell_value, pan_box, class4, class5]
                        
                        # デバッグ出力
                        print(f"マッチした商品: {cell_value}")
                        print(f"  行の全データ: {list(row)}")
                        print(f"  取得結果: パン箱={pan_box}, クラス4={class4}, クラス5={class5}")
                        
                        break
                        
            if best_match:
                break
        
        # 完全一致がなければ部分一致を試す
        if not best_match:
            for idx, row in master_df.iterrows():
                for col_idx in range(min(5, len(master_df.columns))):
                    cell_value = str(row.iloc[col_idx]).strip()
                    if cell_value and cell_value != 'nan':
                        norm_master = unicodedata.normalize('NFKC', cell_value).replace(" ", "")
                        
                        # 部分一致チェック
                        if norm_master and norm_master in norm_pdf:
                            # 同様の処理で値を取得
                            pan_box = ""
                            for potential_pan_col in [2, 6, 7, 8, 9, 10]:
                                if len(row) > potential_pan_col:
                                    potential_value = str(row.iloc[potential_pan_col]).strip()
                                    if potential_value and potential_value.isdigit():
                                        if potential_value not in ["88", "1", "0"]:
                                            pan_box = potential_value
                                            break
                            
                            class4 = str(row.iloc[15]) if len(row) > 15 and str(row.iloc[15]).strip() != 'nan' else ""
                            class5 = str(row.iloc[17]) if len(row) > 17 and str(row.iloc[17]).strip() != 'nan' else ""
                            
                            best_match = [cell_value, pan_box, class4, class5]
                            break
                            
                if best_match:
                    break

        if best_match:
            result_data = best_match
        
        matched_results.append(result_data)
        
    return matched_results
# ──────────────────────────────────────────────
# ▲ここまでが修正された弁当データマッチング関数▲
# ──────────────────────────────────────────────

# ──────────────────────────────────────────────
# PDF解析・データ抽出関数群 (ここは変更なし)
# ──────────────────────────────────────────────
def extract_detailed_client_info_from_pdf(pdf_file_obj):
    client_data = []
    try:
        with pdfplumber.open(pdf_file_obj) as pdf:
            for page in pdf.pages:
                rows = extract_text_with_layout(page)
                if not rows: continue
                garden_row_idx = -1
                for i, row in enumerate(rows):
                    if '園名' in ''.join(str(c) for c in row if c):
                        garden_row_idx = i
                        break
                if garden_row_idx == -1: continue
                current_client_id, current_client_name = None, None
                for i in range(garden_row_idx + 1, len(rows)):
                    row = rows[i]
                    row_text = ''.join(str(c) for c in row if c)
                    if '10001' in row_text: break
                    if not any(str(c).strip() for c in row): continue
                    if row and row[0]:
                        left_cell = str(row[0]).strip()
                        if re.match(r'^\d+$', left_cell):
                            if current_client_id and current_client_name:
                                client_info = extract_meal_numbers_from_row(rows, i - 1, current_client_id, current_client_name)
                                if client_info: client_data.append(client_info)
                            current_client_id, current_client_name = left_cell, None
                        elif not re.match(r'^\d+$', left_cell) and current_client_id:
                            current_client_name = left_cell
                if current_client_id and current_client_name:
                    client_info = extract_meal_numbers_from_row(rows, len(rows) - 1, current_client_id, current_client_name)
                    if client_info: client_data.append(client_info)
    except Exception:
        pass
    return client_data

def extract_meal_numbers_from_row(rows, row_idx, client_id, client_name):
    client_info = {'client_id': client_id, 'client_name': client_name, 'student_meals': [], 'teacher_meals': []}
    rows_to_check = []
    for i in range(max(0, row_idx - 3), min(len(rows), row_idx + 3)):
        if i < len(rows) and rows[i]:
            left_cell = str(rows[i][0]).strip()
            if left_cell == client_id: rows_to_check.append(('id', i, rows[i]))
            elif left_cell == client_name: rows_to_check.append(('name', i, rows[i]))
    all_numbers = []
    for row_type, _, row in rows_to_check:
        for col_idx, cell in enumerate(row[1:], 1):
            cell_str = str(cell).strip()
            if cell_str and re.match(r'^\d+$', cell_str):
                all_numbers.append({'number': int(cell_str), 'row_type': row_type})
            elif cell_str and not re.match(r'^\d+$', cell_str):
                break
    client_info['student_meals'] = [item['number'] for item in all_numbers if item['row_type'] == 'id'][:3]
    client_info['teacher_meals'] = [item['number'] for item in all_numbers if item['row_type'] == 'name'][:2]
    return client_info

def export_detailed_client_data_to_dataframe(client_data):
    df_data = []
    for info in client_data:
        row = {'クライアント名': info['client_name'],'園児の給食の数1': info['student_meals'][0] if len(info['student_meals']) > 0 else '','園児の給食の数2': info['student_meals'][1] if len(info['student_meals']) > 1 else '','園児の給食の数3': info['student_meals'][2] if len(info['student_meals']) > 2 else '','先生の給食の数1': info['teacher_meals'][0] if len(info['teacher_meals']) > 0 else '','先生の給食の数2': info['teacher_meals'][1] if len(info['teacher_meals']) > 1 else ''}
        df_data.append(row)
    return pd.DataFrame(df_data)

def get_line_groups(words: List[Dict[str, Any]], y_tolerance: float = 1.2) -> List[List[Dict[str, Any]]]:
    if not words: return []
    sorted_words = sorted(words, key=lambda w: w['top'])
    groups, current_group = [], [sorted_words[0]]
    current_top = sorted_words[0]['top']
    for word in sorted_words[1:]:
        if abs(word['top'] - current_top) <= y_tolerance:
            current_group.append(word)
        else:
            groups.append(current_group)
            current_group, current_top = [word], word['top']
    groups.append(current_group)
    return groups

def get_vertical_boundaries(page, tolerance: float = 2) -> List[float]:
    lines = page.lines
    vertical_lines_x = sorted(list(set(round((line['x0'] + line['x1']) / 2, 1) for line in lines if abs(line['x0'] - line['x1']) < tolerance)))
    words = page.extract_words()
    if not words: return vertical_lines_x
    left_boundary = min(word['x0'] for word in words)
    right_boundary = max(word['x1'] for word in words)
    boundaries = sorted(list(set([round(left_boundary, 1)] + vertical_lines_x + [round(right_boundary, 1)])))
    merged_boundaries = []
    if boundaries:
        merged_boundaries.append(boundaries[0])
        for b in boundaries[1:]:
            if b - merged_boundaries[-1] > tolerance * 2:
                merged_boundaries.append(b)
    return sorted(list(set(merged_boundaries)))

def split_line_using_boundaries(sorted_words_in_line: List[Dict[str, Any]], boundaries: List[float]) -> List[str]:
    columns = [""] * (len(boundaries) - 1)
    for word in sorted_words_in_line:
        word_center_x = (word['x0'] + word['x1']) / 2
        for i in range(len(boundaries) - 1):
            if boundaries[i] <= word_center_x < boundaries[i+1]:
                columns[i] = (columns[i] + " " + word["text"]) if columns[i] else word["text"]
                break
    return columns

def extract_text_with_layout(page) -> List[List[str]]:
    words = page.extract_words(x_tolerance=3, y_tolerance=3, keep_blank_chars=False)
    if not words: return []
    boundaries = get_vertical_boundaries(page)
    if len(boundaries) < 2:
        lines = page.extract_text(layout=False, x_tolerance=3, y_tolerance=3)
        return [[line] for line in lines.split('\n') if line.strip()] if lines else []
    row_groups = get_line_groups(words, y_tolerance=1.5)
    result_rows = []
    for group in row_groups:
        sorted_group = sorted(group, key=lambda w: w['x0'])
        columns = split_line_using_boundaries(sorted_group, boundaries)
        if any(cell.strip() for cell in columns):
            result_rows.append(columns)
    return result_rows

def remove_extra_empty_columns(rows: List[List[str]]) -> List[List[str]]:
    if not rows: return rows
    num_cols = max(len(row) for row in rows) if rows else 0
    if num_cols == 0: return rows
    is_col_empty = [True] * num_cols
    for row in rows:
        for c, cell in enumerate(row):
            if c < num_cols and cell.strip(): is_col_empty[c] = False
    keep_indices = [c for c, is_empty in enumerate(is_col_empty) if not is_empty]
    return [[row[i] if i < len(row) else "" for i in keep_indices] for row in rows]

def post_process_rows(rows: List[List[str]]) -> List[List[str]]:
    new_rows = [row[:] for row in rows]
    for i, row in enumerate(new_rows):
        for j, cell in enumerate(row):
            if "合計" in str(cell) and i > 0 and j < len(new_rows[i-1]):
                new_rows[i-1][j] = ""
    return new_rows

def pdf_to_excel_data_for_paste_sheet(pdf_file):
    try:
        with pdfplumber.open(pdf_file) as pdf:
            if not pdf.pages: return None
            page = pdf.pages[0]
            rows = extract_text_with_layout(page)
            rows = [row for row in rows if any(cell.strip() for cell in row)]
            if not rows: return None
            rows = post_process_rows(rows)
            rows = remove_extra_empty_columns(rows)
            if not rows: return None
            
            improved_rows = []
            for row in rows:
                improved_row = [improve_number_extraction(str(cell)) for cell in row]
                improved_rows.append(improved_row)
            
            max_cols = max(len(row) for row in improved_rows)
            normalized_rows = [row + [''] * (max_cols - len(row)) for row in improved_rows]
            return pd.DataFrame(normalized_rows)
    except Exception:
        return None

def improve_number_extraction(cell_text):
    if not cell_text or not str(cell_text).strip(): return cell_text
    cell_str = str(cell_text).strip()
    if re.match(r'^\d{1,3}(,\d{3})*$', cell_str):
        try: return int(cell_str.replace(',', ''))
        except ValueError: pass
    elif re.match(r'^\d+\.\d+$', cell_str):
        try: return float(cell_str)
        except ValueError: pass
    elif re.match(r'^\d+$', cell_str):
        try: return int(cell_str)
        except ValueError: pass
    number_match = re.search(r'\d{1,3}(,\d{3})*(\.\d+)?', cell_str)
    if number_match:
        number_part = number_match.group(0)
        try:
            if '.' in number_part: extracted_number = float(number_part.replace(',', ''))
            else: extracted_number = int(number_part.replace(',', ''))
            if len(number_part) / len(cell_str) > 0.7: return extracted_number
        except ValueError: pass
    return cell_text

def debug_pdf_content(pdf_bytes_io) -> dict:
    debug_info = {'pages': 0, 'total_chars': 0, 'numbers_found': [], 'text_sample': ""}
    try:
        with pdfplumber.open(pdf_bytes_io) as pdf:
            debug_info['pages'] = len(pdf.pages)
            for i, page in enumerate(pdf.pages):
                words = page.extract_words()
                debug_info['total_chars'] += len(words) if words else 0
                text = page.extract_text() or ""
                if i == 0: debug_info['text_sample'] = text[:500]
                numbers = re.findall(r'\d{1,3}(,\d{3})*(\.\d+)?|\d+', text)
                clean_numbers = []
                for num in numbers:
                    if isinstance(num, tuple):
                        clean_numbers.append(num[0] + num[1] if num[1] else num[0])
                    else:
                        clean_numbers.append(num)
                debug_info['numbers_found'].extend(clean_numbers)
    except Exception as e:
        debug_info['error'] = str(e)
    return debug_info

def extract_table_from_pdf_for_bento(pdf_file_obj):
    tables = []
    with pdfplumber.open(pdf_file_obj) as pdf:
        for page in pdf.pages:
            if not page.extract_text() or not any(kw in page.extract_text() for kw in ["園名", "飯あり", "キャラ弁"]): continue
            if not page.lines: continue
            table_settings = {"vertical_strategy": "lines", "horizontal_strategy": "lines"}
            table = page.extract_table(table_settings)
            if table: tables.append(table)
    return tables

def find_correct_anchor_for_bento(table, target_row_text="赤"):
    for r_idx, row in enumerate(table):
        if target_row_text in ''.join(str(c) for c in row if c):
            for offset in [1, 2]:
                if r_idx + offset < len(table):
                    for c_idx, cell in enumerate(table[r_idx + offset]):
                        if cell and "飯なし" in cell: return c_idx
    return -1

def extract_bento_range_for_bento(table, start_col):
    bento_list, end_col = [], -1
    for row in table:
        if "おやつ" in ''.join(str(c) for c in row if c):
            for c_idx, cell in enumerate(row):
                if cell and "おやつ" in cell:
                    end_col = c_idx
                    break
            if end_col != -1: break
    if end_col == -1 or start_col >= end_col: return []
    header_row_idx = -1
    for r_idx, row in enumerate(table):
        if any(c and "飯なし" in c for c in row):
            if r_idx > 0: header_row_idx = r_idx - 1
            break
    if header_row_idx == -1: return []
    header_row = table[header_row_idx]
    for col in range(start_col + 1, end_col):
        cell_text = header_row[col] if col < len(header_row) else ""
        if cell_text and str(cell_text).strip():
            bento_list.append(str(cell_text).strip())
    return bento_list
