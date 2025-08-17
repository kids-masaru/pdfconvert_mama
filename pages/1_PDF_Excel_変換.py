import streamlit as st
import pandas as pd
import io
import os
import re
from openpyxl import load_workbook
import traceback

# 以前の長い関数は、utilsファイルから読み込む
from pdf_utils import (
    safe_write_df,
    pdf_to_excel_data_for_paste_sheet,
    extract_table_from_pdf_for_bento,
    find_correct_anchor_for_bento,
    extract_bento_range_for_bento,
    match_bento_names,
    extract_detailed_client_info_from_pdf,
    export_detailed_client_data_to_dataframe,
)

# ページのタイトル
st.markdown('<div class="title">【数出表】PDF → Excelへの変換</div>', unsafe_allow_html=True)
st.markdown("---")

# PDFアップローダー
uploaded_pdf = st.file_uploader("処理するPDFファイルをアップロードしてください", type="pdf")

if uploaded_pdf is not None:
    # テンプレートファイルの存在チェック
    template_path = "template.xlsm"
    nouhinsyo_path = "nouhinsyo.xlsx"
    if not os.path.exists(template_path) or not os.path.exists(nouhinsyo_path):
        st.error(f"'{template_path}' または '{nouhinsyo_path}' が見つかりません。")
        st.stop()
    
    # テンプレートをメモリに読み込む
    template_wb = load_workbook(template_path, keep_vba=True)
    nouhinsyo_wb = load_workbook(nouhinsyo_path)

    pdf_bytes_io = io.BytesIO(uploaded_pdf.getvalue())
    
    # データ抽出処理
    df_paste_sheet, df_bento_sheet, df_client_sheet = None, None, None
    with st.spinner("PDFからデータを抽出中..."):
        df_paste_sheet = pdf_to_excel_data_for_paste_sheet(io.BytesIO(pdf_bytes_io.getvalue()))
        if df_paste_sheet is not None:
            try:
                tables = extract_table_from_pdf_for_bento(io.BytesIO(pdf_bytes_io.getvalue()))
                if tables:
                    main_table = max(tables, key=len)
                    anchor_col = find_correct_anchor_for_bento(main_table)
                    if anchor_col != -1:
                        bento_list = extract_bento_range_for_bento(main_table, anchor_col)
                        if bento_list:
                            matched_list = match_bento_names(bento_list, st.session_state.master_df)
                            output_data = []
                            for item in matched_list:
                                match = re.search(r' \(入数: (.+?)\)$', item)
                                if match:
                                    output_data.append([item[:match.start()], match.group(1)])
                                else:
                                    output_data.append([item.replace(" (未マッチ)", ""), ""])
                            df_bento_sheet = pd.DataFrame(output_data, columns=['商品予定名', 'パン箱入数'])
            except Exception as e: st.error(f"注文弁当データ処理中にエラー: {e}")
            
            try:
                client_data = extract_detailed_client_info_from_pdf(io.BytesIO(pdf_bytes_io.getvalue()))
                if client_data:
                    df_client_sheet = export_detailed_client_data_to_dataframe(client_data)
                    st.success(f"クライアント情報 {len(client_data)} 件を抽出しました")
            except Exception as e: st.error(f"クライアント情報抽出中にエラー: {e}")
    
    # Excelファイルへの書き込みとダウンロード
    if df_paste_sheet is not None:
        try:
            with st.spinner("Excelファイルを作成中..."):
                # --- template.xlsmへの書き込み ---
                ws_paste = template_wb["貼り付け用"]
                for r_idx, row in df_paste_sheet.iterrows():
                    for c_idx, value in enumerate(row):
                        ws_paste.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                if df_bento_sheet is not None: safe_write_df(template_wb["注文弁当の抽出"], df_bento_sheet, start_row=1)
                if df_client_sheet is not None: safe_write_df(template_wb["クライアント抽出"], df_client_sheet, start_row=1)
                output_macro = io.BytesIO()
                template_wb.save(output_macro)
                macro_excel_bytes = output_macro.getvalue()

                # --- nouhinsyo.xlsxへの書き込み ---
                df_bento_for_nouhin = None
                if df_bento_sheet is not None:
                    master_df = st.session_state.master_df
                    master_map = master_df.drop_duplicates(subset=['商品予定名']).set_index('商品予定名')['商品名'].to_dict()
                    df_bento_for_nouhin = df_bento_sheet.copy()
                    df_bento_for_nouhin['商品名'] = df_bento_for_nouhin['商品予定名'].map(master_map)
                    df_bento_for_nouhin = df_bento_for_nouhin[['商品予定名', 'パン箱入数', '商品名']]
                ws_paste_n = nouhinsyo_wb["貼り付け用"]
                for r_idx, row in df_paste_sheet.iterrows():
                    for c_idx, value in enumerate(row):
                        ws_paste_n.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                if df_bento_for_nouhin is not None: safe_write_df(nouhinsyo_wb["注文弁当の抽出"], df_bento_for_nouhin, start_row=1)
                if df_client_sheet is not None: safe_write_df(nouhinsyo_wb["クライアント抽出"], df_client_sheet, start_row=1)
                output_data_only = io.BytesIO()
                nouhinsyo_wb.save(output_data_only)
                data_only_excel_bytes = output_data_only.getvalue()

            st.success("✅ ファイルの準備が完了しました！")
            original_pdf_name = os.path.splitext(uploaded_pdf.name)[0]
            
            col1, col2 = st.columns(2)
            with col1:
                st.download_button(label="【数出表ダウンロード】",data=macro_excel_bytes,file_name=f"{original_pdf_name}_Processed.xlsm",mime="application/vnd.ms-excel.sheet.macroEnabled.12")
            with col2:
                st.download_button(label="【納品書ダウンロード】",data=data_only_excel_bytes,file_name=f"{original_pdf_name}_Data.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            st.error(f"Excelファイル生成中にエラーが発生しました: {e}")
            traceback.print_exc()
