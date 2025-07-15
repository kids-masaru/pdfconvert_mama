import streamlit as st
import pdfplumber
import pandas as pd
import io
import re
import base64
import os
import unicodedata
import traceback
from typing import List, Dict, Any
from openpyxl import load_workbook

# テンプレートExcelファイルのパス設定と存在確認 (各ページで共通のパスを使用)
script_dir = os.path.dirname(os.path.abspath(__file__))
# pagesフォルダから見て一つ上の階層にあることを想定
template_path = os.path.join(script_dir, "..", "template.xlsm")
master_csv_path = os.path.join(script_dir, "..", "商品マスタ一覧.csv")

# セッションステートにテンプレートワークブックを格納（一度だけ読み込む）
# アプリケーション全体でtemplate_wbを共有するために、streamlit_app.pyで読み込み、
# session_state経由で渡すのが理想的ですが、簡略化のため各ページでチェック
# 通常はstreamlit_app.pyで読み込み、st.session_stateに保存し、各ページではそれを参照します。
# ここでは、PDF処理ページのみでtemplate_wbが必要なため、このページで読み込みを管理します。
if 'template_wb' not in st.session_state:
    st.session_state.template_wb = None

if not st.session_state.template_wb:
    if not os.path.exists(template_path):
        st.error(f"エラー: テンプレートファイル '{template_path}' が見つかりません。スクリプトの隣に配置してください。")
        st.stop() # アプリの実行を停止
    if not os.path.exists(master_csv_path):
        st.error(f"エラー: 商品マスタファイル '{master_csv_path}' が見つかりません。スクリプトの隣に配置してください。")
        st.stop() # アプリの実行を停止
    
    try:
        st.session_state.template_wb = load_workbook(template_path, keep_vba=True)
        # st.success("テンプレートファイルを読み込みました。") # デバッグ用
    except Exception as e:
        st.error(f"エラー: テンプレートファイル '{template_path}' の読み込み中にエラーが発生しました: {e}")
        st.session_state.template_wb = None
        st.stop() # アプリの実行を停止

template_wb = st.session_state.template_wb


# ──────────────────────────────────────────────
# PDF→Excel 変換ロジック (Streamlit版から)
# ──────────────────────────────────────────────
# この部分は以前のpdfconvert_mama.pyから変更なしでコピーします
def is_number(text: str) -> bool:
    return bool(re.match(r'^\d+$', text.strip()))

def get_line_groups(words: List[Dict[str, Any]], y_tolerance: float = 1.2) -> List[List[Dict[str, Any]]]:
    """y座標に基づいて単語を行ごとにグループ化する"""
    if not words:
        return []
    sorted_words = sorted(words, key=lambda w: w['top'])
    groups = []
    current_group = [sorted_words[0]]
    current_top = sorted_words[0]['top']
    for word in sorted_words[1:]:
        if abs(word['top'] - current_top) <= y_tolerance:
            current_group.append(word)
        else:
            groups.append(current_group)
            current_group = [word]
            current_top = word['top']
    groups.append(current_group)
    return groups

def get_vertical_boundaries(page, tolerance: float = 2) -> List[float]:
    """ページの縦線と単語の左右端から列の境界線を推定する"""
    vertical_lines_x = []
    for line in page.lines:
        if abs(line['x0'] - line['x1']) < tolerance:
            vertical_lines_x.append((line['x0'] + line['x1']) / 2)
    vertical_lines_x = sorted(list(set(round(x, 1) for x in vertical_lines_x)))

    words = page.extract_words()
    if not words:
        return vertical_lines_x

    left_boundary = min(word['x0'] for word in words)
    right_boundary = max(word['x1'] for word in words)

    boundaries = sorted(list(set([round(left_boundary, 1)] + vertical_lines_x + [round(right_boundary, 1)])))

    merged_boundaries = []
    if boundaries:
        merged_boundaries.append(boundaries[0])
        for i in range(1, len(boundaries)):
            if boundaries[i] - merged_boundaries[-1] > tolerance * 2:
                merged_boundaries.append(boundaries[i])
        if right_boundary > merged_boundaries[-1] + tolerance * 2 :
                merged_boundaries.append(round(right_boundary, 1))
        boundaries = sorted(list(set(merged_boundaries)))

    return boundaries

def split_line_using_boundaries(sorted_words_in_line: List[Dict[str, Any]], boundaries: List[float]) -> List[str]:
    """境界線に基づいて、一行分の単語をセルテキストに分割する"""
    columns = [""] * (len(boundaries) - 1)
    for word in sorted_words_in_line:
        word_center_x = (word['x0'] + word['x1']) / 2
        for i in range(len(boundaries) - 1):
            left = boundaries[i]
            right = boundaries[i + 1]
            if left <= word_center_x < right:
                if columns[i]:
                    columns[i] += " " + word["text"]
                else:
                    columns[i] = word["text"]
                break
    return columns

def extract_text_with_layout(page) -> List[List[str]]:
    """PDFページからレイアウトを考慮してテキストを行と列に抽出する"""
    words = page.extract_words(x_tolerance=3, y_tolerance=3, keep_blank_chars=False)
    if not words:
        return []

    boundaries = get_vertical_boundaries(page)
    if len(boundaries) < 2:
            lines = page.extract_text(layout=False, x_tolerance=3, y_tolerance=3)
            return [[line] for line in lines.split('\n') if line.strip()]

    row_groups = get_line_groups(words, y_tolerance=1.5)

    result_rows = []
    for group in row_groups:
        sorted_group = sorted(group, key=lambda w: w['x0'])
        columns = split_line_using_boundaries(sorted_group, boundaries)
        if any(cell.strip() for cell in columns):
            result_rows.append(columns)

    return result_rows

def remove_extra_empty_columns(rows: List[List[str]]) -> List[List[str]]:
    """すべての行で完全に空である列を削除する"""
    if not rows:
        return rows
    num_cols = max(len(row) for row in rows) if rows else 0
    if num_cols == 0:
        return rows

    is_col_empty = [True] * num_cols
    for r, row in enumerate(rows):
        for c in range(len(row)):
            if c < num_cols and row[c].strip():
                is_col_empty[c] = False

    keep_indices = [c for c in range(num_cols) if not is_col_empty[c]]

    new_rows = []
    for row in rows:
        new_row = [row[i] if i < len(row) else "" for i in keep_indices]
        new_rows.append(new_row)

    return new_rows

def post_process_rows(rows: List[List[str]]) -> List[List[str]]:
    """データの後処理: 例として「合計」行の上のセルをクリア"""
    new_rows = [row[:] for row in rows]
    for i, row in enumerate(new_rows):
        for j, cell in enumerate(row):
            if "合計" in str(cell):
                if i > 0 and j < len(new_rows[i-1]):
                    new_rows[i-1][j] = ""
    return new_rows

def pdf_to_excel_data_for_paste_sheet(pdf_file) -> pd.DataFrame | None:
    """
    PDFファイルを読み込み、最初のページの表形式データをpandas DataFrameとして返す。
    「貼り付け用」シート向け。
    """
    try:
        # pdfplumber.open はファイルパスまたはバイナリI/Oオブジェクトを受け取る
        with pdfplumber.open(pdf_file) as pdf:
            if not pdf.pages:
                st.warning("PDFにページがありません。")
                return None
            page = pdf.pages[0]

            rows = extract_text_with_layout(page)
            rows = [row for row in rows if any(cell.strip() for cell in row)]
            if not rows:
                st.warning("PDFの最初のページからテキストデータを抽出できませんでした。（貼り付け用）")
                return None

            rows = post_process_rows(rows)
            rows = remove_extra_empty_columns(rows)
            if not rows or not rows[0]:
                    st.warning("空の列を削除した結果、データがなくなりました。（貼り付け用）")
                    return None

            max_cols = max(len(row) for row in rows) if rows else 0
            normalized_rows = [row + [''] * (max_cols - len(row)) for row in rows]
            df = pd.DataFrame(normalized_rows)
            return df

    except Exception as e:
        st.error(f"PDF処理中にエラーが発生しました（貼り付け用）: {e}")
        return None

# ──────────────────────────────────────────────
# PDF→Excel 変換ロジック (CLI版から)
# ──────────────────────────────────────────────
# この部分も以前のpdfconvert_mama.pyから変更なしでコピーします
def extract_table_from_pdf_for_bento(pdf_file_obj):
    """PDFから線で囲まれた表領域を正確に抽出 (「注文弁当の抽出」用)"""
    tables = []
    # pdfplumber.open はファイルパスまたはバイナリI/Oオブジェクトを受け取る
    with pdfplumber.open(pdf_file_obj) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            
            start_keywords = ["園名", "飯あり", "キャラ弁"]
            end_keywords = ["おやつ", "合計", "PAGE"]
            
            if not any(kw in text for kw in start_keywords):
                continue
                
            lines = page.lines
            if not lines:
                continue
                
            y_coords = sorted(set([line['top'] for line in lines] + [line['bottom'] for line in lines]))
            if len(y_coords) < 2:
                continue
                
            table_top = min(y_coords)
            table_bottom = max(y_coords)
            
            x_coords = sorted(set([line['x0'] for line in lines] + [line['x1'] for line in lines]))
            if len(x_coords) < 2:
                continue
                
            table_left = min(x_coords)
            table_right = max(x_coords)
            
            table_bbox = (table_left, table_top, table_right, table_bottom)
            cropped_page = page.crop(table_bbox)
            
            table_settings = {
                "vertical_strategy": "lines",
                "horizontal_strategy": "lines",
                "snap_tolerance": 3,
                "join_tolerance": 3,
                "edge_min_length": 15,
            }
            
            table = cropped_page.extract_table(table_settings)
            if table:
                tables.append(table)
    
    return tables

def find_correct_anchor_for_bento(table, target_row_text="赤"):
    """「赤」行の直下にある「飯なし」を特定 (「注文弁当の抽出」用)"""
    for row_idx, row in enumerate(table):
        row_text = ''.join(str(cell) for cell in row if cell)
        if target_row_text in row_text:
            for offset in [1, 2]:
                if row_idx + offset < len(table):
                    next_row = table[row_idx + offset]
                    for col_idx, cell in enumerate(next_row):
                        if cell and "飯なし" in cell:
                            return col_idx
    return -1

def extract_bento_range_for_bento(table, start_col):
    """「飯なし」から「おやつ」までの範囲を抽出 (「注文弁当の抽出」用)"""
    bento_list = []
    end_col = -1
    
    for row in table:
        row_text = ''.join(str(cell) for cell in row if cell)
        if "おやつ" in row_text:
            for col_idx, cell in enumerate(row):
                if cell and "おやつ" in cell:
                    end_col = col_idx
                    break
            if end_col != -1:
                break
    
    if end_col == -1 or start_col >= end_col:
        return []
    
    header_row_idx = None
    anchor_row_idx = -1
    for row_idx, row in enumerate(table):
        if any(cell and "飯なし" in cell for cell in row):
            anchor_row_idx = row_idx
            break
    
    if anchor_row_idx == -1:
        return []
    
    if anchor_row_idx - 1 >= 0:
        header_row_idx = anchor_row_idx - 1
    else:
        return []
    
    for col in range(start_col + 1, end_col + 1):
        if col < len(table[header_row_idx]):
            cell_text = table[header_row_idx][col]
        else:
            cell_text = ""
        
        if cell_text and str(cell_text).strip() and "飯なし" not in str(cell_text):
            bento_list.append(str(cell_text).strip())
    
    return bento_list

def match_bento_names(pdf_bento_list, master_csv_path):
    """
    マスタデータと部分一致で照合し、I列の数字も一緒に表示
    さらに、未マッチの場合にPDF名を右端から削って再照合する。
    """
    encodings = ['utf-8', 'shift_jis', 'cp932', 'euc-jp', 'iso-2022-jp']
    
    master_df = None
    
    for encoding in encodings:
        try:
            temp_df = pd.read_csv(master_csv_path, encoding=encoding)
            
            if not temp_df.empty:
                master_df = temp_df
                break
            else:
                pass

        except UnicodeDecodeError:
            continue
        except Exception as e:
            continue
    else:
        st.error("マスタデータの読み込みに失敗しました。サポートされているエンコーディングを試しましたが、どれも機能しませんでした。")
        return []
    
    if master_df is None:
        return []

    master_data_tuples = []
    
    try:
        if '商品予定名' in master_df.columns and 'パン箱入数' in master_df.columns:
            master_data_tuples = master_df[['商品予定名', 'パン箱入数']].dropna().values.tolist()
            master_data_tuples = [(str(name), str(value)) for name, value in master_data_tuples]
        elif '商品予定名' in master_df.columns:
            st.warning("警告: マスタCSVに「パン箱入数」列が見つかりません。商品予定名のみで照合します。")
            master_data_tuples = master_df['商品予定名'].dropna().astype(str).tolist()
            master_data_tuples = [(name, "") for name in master_data_tuples]
        else:
            st.error("エラー: マスタCSVに「商品予定名」列が見つかりません。")
            return []

    except KeyError as e:
        st.error(f"エラー: マスタCSVに必要な列が見つかりません: {e}。CSVのヘッダー名を確認してください。")
        return []
    except Exception as e:
        st.error(f"マスタデータ処理中に予期せぬエラーが発生しました: {e}")
        return []
    
    if len(master_data_tuples) == 0:
        st.warning("マスタデータから有効な商品情報が抽出できませんでした。")
        return []

    matched = []
    
    normalized_master_data_tuples = []
    for master_name, master_id in master_data_tuples:
        normalized_name = unicodedata.normalize('NFKC', master_name)
        normalized_name = re.sub(r'\s+', '', normalized_name)
        normalized_master_data_tuples.append((normalized_name, master_name, master_id))
    
    for pdf_name in pdf_bento_list:
        original_normalized_pdf_name = unicodedata.normalize('NFKC', str(pdf_name))
        original_normalized_pdf_name = re.sub(r'\s+', '', original_normalized_pdf_name)
        
        current_pdf_name_for_matching = original_normalized_pdf_name
        
        found_match = False
        found_original_master_name = None
        found_id = None
        
        for norm_m_name, orig_m_name, m_id in normalized_master_data_tuples:
            if norm_m_name.startswith(current_pdf_name_for_matching):
                found_original_master_name = orig_m_name
                found_id = m_id
                found_match = True
                break
        
        if not found_match:
            for norm_m_name, orig_m_name, m_id in normalized_master_data_tuples:
                if current_pdf_name_for_matching in norm_m_name:
                    found_original_master_name = orig_m_name
                    found_id = m_id
                    found_match = True
                    break
        
        if not found_match:
            for num_chars_to_remove in range(1, 4): 
                if len(original_normalized_pdf_name) > num_chars_to_remove:
                    truncated_pdf_name = original_normalized_pdf_name[:-num_chars_to_remove]
                    
                    for norm_m_name, orig_m_name, m_id in normalized_master_data_tuples:
                        if norm_m_name.startswith(truncated_pdf_name):
                            found_original_master_name = orig_m_name
                            found_id = m_id
                            found_match = True
                            break
                    
                    if not found_match:
                        for norm_m_name, orig_m_name, m_id in normalized_master_data_tuples:
                            if truncated_pdf_name in norm_m_name:
                                found_original_master_name = orig_m_name
                                found_id = m_id
                                found_match = True
                                break
                
                if found_match:
                    break
        
        if found_original_master_name:
            if found_id:
                matched.append(f"{found_original_master_name} (入数: {found_id})")
            else:
                matched.append(found_original_master_name)
        else:
            matched.append(f"{pdf_name} (未マッチ)")
    
    return matched


st.markdown('<div class="main-container">', unsafe_allow_html=True)
st.markdown('<div class="title">PDF → Excel 変換</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">PDFの注文データをアップロードし、Excelファイルに変換します。</div>', unsafe_allow_html=True)

# ----------------------------
# UI：PDFファイルアップロード
# ----------------------------
uploaded_pdf = st.file_uploader("処理するPDFファイルをアップロードしてください", type="pdf",
                                 help="ここにPDFファイルをドラッグ＆ドロップするか、クリックして選択してください。")

# --- ファイル処理とダウンロード表示用のコンテナ ---
file_container = st.container()
download_container = st.container()

# --- PDFがアップロードされたら処理を実行 ---
if uploaded_pdf is not None and template_wb is not None:
    # --- 処理中の表示 ---
    with file_container:
        file_ext = uploaded_pdf.name.split('.')[-1].lower()
        file_icon = "PDF"
        file_size = len(uploaded_pdf.getvalue()) / 1024

        progress_placeholder = st.empty()
        progress_placeholder.markdown(f"""
        <div class="file-card">
            <div class="file-info">
                <div class="file-icon">{file_icon}</div>
                <div class="file-details">
                    <div class="file-name">{uploaded_pdf.name}</div>
                    <div class="file-meta">{file_size:.0f} KB</div>
                </div>
            </div>
            <div class="loading-spinner"></div>
        </div>
        <div class="progress-bar"><div class="progress-value"></div></div>
        """, unsafe_allow_html=True)

    # --- PDFのバイナリデータをio.BytesIOに変換 (pdfplumberが直接処理できるように) ---
    pdf_bytes_io = io.BytesIO(uploaded_pdf.getvalue())


    # --- DataFrameへの変換（貼り付け用シート向け）---
    df_paste_sheet = None
    with st.spinner("「貼り付け用」データを抽出中..."):
        pdf_bytes_io.seek(0) 
        df_paste_sheet = pdf_to_excel_data_for_paste_sheet(pdf_bytes_io)

    # --- DataFrameへの変換（注文弁当の抽出シート向け）---
    df_bento_sheet = None
    if df_paste_sheet is not None: # 貼り付け用データが成功した場合のみ次へ
        with st.spinner("「注文弁当の抽出」データを抽出中..."):
            try:
                pdf_bytes_io.seek(0) # ポインタをリセット
                tables = extract_table_from_pdf_for_bento(pdf_bytes_io)
                if not tables:
                    st.warning("PDFから表を抽出できませんでした。（注文弁当の抽出）")
                else:
                    main_table = max(tables, key=lambda t: len(t) * len(t[0])) if tables else []
                    if not main_table:
                        st.warning("メインとなる表が見つかりませんでした。（注文弁当の抽出）")
                    else:
                        anchor_col = find_correct_anchor_for_bento(main_table)
                        if anchor_col == -1:
                            st.warning("「赤」行下の「飯なし」を見つけられませんでした。（注文弁当の抽出）")
                        else:
                            bento_list = extract_bento_range_for_bento(main_table, anchor_col)
                            if not bento_list:
                                st.warning("弁当範囲を抽出できませんでした。（注文弁当の抽出）")
                            else:
                                matched_list = match_bento_names(bento_list, master_csv_path)
                                output_data_bento = []
                                for item in matched_list:
                                    match_found = False
                                    match = re.search(r' \(入数: (.+?)\)$', item)
                                    if match:
                                        bento_name = item[:match.start()]
                                        bento_count = match.group(1)
                                        output_data_bento.append([bento_name.strip(), bento_count.strip()])
                                        match_found = True
                                    elif "(未マッチ)" in item:
                                        bento_name = item.replace(" (未マッチ)", "").strip()
                                        bento_count = ""
                                        output_data_bento.append([bento_name, bento_count])
                                        match_found = True
                                    if not match_found:
                                        output_data_bento.append([item.strip(), ""])
                                df_bento_sheet = pd.DataFrame(output_data_bento, columns=['商品予定名', 'パン箱入数'])
            except Exception as e:
                st.error(f"「注文弁当の抽出」データ処理中にエラーが発生しました: {e}")
                st.exception(e)

    # --- Excelに書き込み ---
    if df_paste_sheet is not None and (df_bento_sheet is not None or not (tables and main_table and bento_list)):
        try:
            with st.spinner("Excelテンプレートにデータを書き込み中..."):
                # 「貼り付け用」シートへの書き込み
                try:
                    ws_paste = template_wb["貼り付け用"]
                    # 既存のデータをクリア (必要であれば)
                    # ws_paste.delete_rows(1, ws_paste.max_row)

                    for r_idx, row in df_paste_sheet.iterrows():
                        for c_idx, value in enumerate(row):
                            ws_paste.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                except KeyError:
                    st.error("エラー: テンプレートファイルに「貼り付け用」という名前のシートが見つかりません。")
                    st.stop()
                
                # 「注文弁当の抽出」シートへの書き込み (df_bento_sheetがNoneでない場合のみ)
                if df_bento_sheet is not None and not df_bento_sheet.empty:
                    try:
                        ws_bento = template_wb["注文弁当の抽出"]
                        # 既存のデータをクリア (必要であれば)
                        # ws_bento.delete_rows(1, ws_bento.max_row)

                        for r_idx, row in df_bento_sheet.iterrows():
                            for c_idx, value in enumerate(row):
                                ws_bento.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                    except KeyError:
                        st.error("エラー: テンプレートファイルに「注文弁当の抽出」という名前のシートが見つかりません。")
                        st.stop()
                elif df_bento_sheet is not None and df_bento_sheet.empty:
                    st.warning("「注文弁当の抽出」シートに書き込むデータがありませんでした。")
                else:
                    st.warning("「注文弁当の抽出」データの準備ができませんでした。このシートへの書き込みはスキップされます。")


            # --- メモリ上でExcelファイルを生成 ---
            with st.spinner("Excelファイルを生成中..."):
                output = io.BytesIO()
                template_wb.save(output)
                output.seek(0)
                final_excel_bytes = output.read()

            # --- 処理完了表示 ---
            with file_container:
                    progress_placeholder.markdown(f"""
                    <div class="file-card">
                        <div class="file-info">
                            <div class="file-icon">{file_icon}</div>
                            <div class="file-details">
                                <div class="file-name">{uploaded_pdf.name}</div>
                                <div class="file-meta">{file_size:.0f} KB - 処理完了</div>
                            </div>
                        </div>
                        <div class="check-icon">✓</div>
                    </div>
                    """, unsafe_allow_html=True)

            # --- ダウンロードリンクの生成 ---
            with download_container:
                st.markdown('<div class="separator"></div>', unsafe_allow_html=True)

                original_pdf_name = os.path.splitext(uploaded_pdf.name)[0]
                output_filename = f"{original_pdf_name}_Processed.xlsm"
                excel_size = len(final_excel_bytes) / 1024
                b64 = base64.b64encode(final_excel_bytes).decode('utf-8')

                mime_type = "application/vnd.ms-excel.sheet.macroEnabled.12"

                href = f"""
                <a href="data:{mime_type};base64,{b64}" download="{output_filename}" class="download-card">
                    <div class="download-info">
                        <div class="download-icon">XLSM</div>
                        <div class="download-details">
                            <div class="download-name">{output_filename}</div>
                            <div class="download-meta">Excel (マクロ有効)・{excel_size:.0f} KB</div>
                        </div>
                    </div>
                    <div class="download-button-imitation">
                        <span class="download-button-icon">↓</span>
                        Download
                    </div>
                </a>
                """
                st.markdown(href, unsafe_allow_html=True)

        except Exception as e:
            st.error(f"Excelファイルへの書き込みまたは生成中にエラーが発生しました: {e}")
            st.exception(e)
            with file_container:
                    progress_placeholder.markdown(f"""
                    <div class="file-card" style="border-color: red;">
                        <div class="file-info">
                            <div class="file-icon" style="background-color: red;">!</div>
                            <div class="file-details">
                                <div class="file-name">{uploaded_pdf.name}</div>
                                <div class="file-meta" style="color: red;">処理中にエラーが発生しました</div>
                            </div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

    else:
        st.warning("PDFデータ抽出に問題があったため、Excelファイルは生成されませんでした。エラーメッセージを確認してください。")
        with file_container:
            progress_placeholder.markdown(f"""
            <div class="file-card" style="border-color: orange;">
                <div class="file-info">
                    <div class="file-icon" style="background-color: orange;">!</div>
                    <div class="file-details">
                        <div class="file-name">{uploaded_pdf.name}</div>
                        <div class="file-meta" style="color: orange;">データ抽出に失敗しました</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)


# --- テンプレートファイルが見つからないか読み込めなかった場合 ---
elif uploaded_pdf is not None and template_wb is None:
    st.warning("テンプレートファイルが正しく読み込めていないため、処理を開始できません。")


st.markdown('</div>', unsafe_allow_html=True)
